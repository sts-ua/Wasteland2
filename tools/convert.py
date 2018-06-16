#!/usr/bin/env python

from openpyxl import Workbook, load_workbook
from sys import argv


def to_xlsx(name: str) -> None:
    print(f'Working with {name}.txt…')
    wb = Workbook()
    ws = wb.active

    with open(f'{name}.txt', encoding='utf-16') as file:
        content = file.readlines()

    def parce_lines(lines, prefix):
        return [t[1:-1] for t in lines
                if t.startswith(prefix) and len(t)]

    keys = parce_lines(content, '#')
    strings = parce_lines(content, '=')

    translation = zip(keys, strings)

    for i, (key, string) in enumerate(translation):
        ws[f'A{i+1}'] = key
        ws[f'B{i+1}'] = string

    wb.save(f'{name}.xlsx')
    print(f'Converted to {name}.xlsx!')


def to_txt(name: str) -> None:
    print(f'Working with {name}.xlsx…')
    wb = load_workbook(f'{name}.xlsx')
    ws = wb.active

    # if there are more than 2 columns -- use third one, else -- second one
    trans_index = 1 if len(ws[1]) == 2 else 2

    translation = [(row[0].value, row[trans_index].value) for row in ws]

    with open(f'{name}.txt', 'w', encoding='utf-16') as file:
        for key, string in translation:
            if {key}.startswith('Identification (key)'):
                continue
            file.write(f'#{key}\n')
            file.write(f'={string}\n\n')

    print(f'Converted to {name}.txt!')


def main(argv):
    for arg in argv[1:]:
        *name, ext = arg.split('.')
        if ext == 'txt':
            to_xlsx('.'.join(name))
        elif ext == 'xlsx':
            to_txt('.'.join(name))
        else:
            print('Wrong format!')


if __name__ == '__main__':
    main(argv)
    input('Press ENTER to close the window…')
